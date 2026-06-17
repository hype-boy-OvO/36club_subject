import torch
import torch.optim as optim
import torch.nn as nn
import torch.nn.functional as F
import models
import pre_process
import pre_process

import os
import openpyxl


model_name = "kykim/bert-kor-base"
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
model = models.My_Model(model_name)
model.to(device)
excel_path = r"datasets\2025_2026_Trend_And_General_Sentences.xlsx"

trendy_sent = pre_process.get_dataloader(excel_path=excel_path,target_category="유행어/신조어")
it_sent = pre_process.get_dataloader(excel_path=excel_path,target_category="IT 신기술")
bio_sent = pre_process.get_dataloader(excel_path=excel_path,target_category="생물/바이오 분야")

model.load_state_dict(
    torch.load("./pretrained_model.pt")
)

graph_encoder_params = list(
    model.graph_memory.graph_encoder.parameters()
)

other_graph_memory_params = [
    p
    for name, p in model.graph_memory.named_parameters()
    if not name.startswith("graph_encoder.")
]

optimizer = optim.AdamW([
    {
        "params": graph_encoder_params,
        "lr": 3e-5
    },
    {
        "params": other_graph_memory_params,
        "lr": 1e-4
    },
    {
        "params": model.prediction_model.parameters(),
        "lr": 3e-5
    },
    {
        "params": model.final_model.parameters(),
        "lr": 3e-5
    }
], weight_decay=1e-4)

file_name = r"history\loss_history.xlsx"
if os.path.exists(file_name):
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
else:
    # 파일이 없으면 새로 만들기
    wb = openpyxl.Workbook()
    ws = wb.active

ws["A1"] = "trend_1"
ws["B1"] = "trend_2"
ws["C1"] = "it_1"
ws["D1"] = "it_2"
ws["E1"] = "bio_1"
ws["F1"] = "bio_2"
ws["M1"] = "gang_1"
ws["N1"] = "hoon_1"
ws["Q1"] = "norm_gang_1"
ws["R1"] = "norm_hoon_1"

ws["G1"] = "norm_trend_1"
ws["H1"] = "norm_trend_2"
ws["I1"] = "norm_it_1"
ws["J1"] = "norm_it_2"
ws["K1"] = "norm_bio_1"
ws["L1"] = "norm_bio_2"
ws["O1"] = "gang_2"
ws["P1"] = "hoon_2"
ws["S1"] = "norm_gang_2"
ws["T1"] = "norm_hoon_2"

def study(dataloader, epochs=70,kind=None):
    next_row = 2
    total_loss = 0
    for epoch in range(epochs):
        for step, batch in enumerate(dataloader):
            batch = {k: v.to(device) for k, v in batch.items()}
        
            optimizer.zero_grad()

            ouput, loss, new_nodes, new_edges = model(**batch)
        
            loss.backward()
            optimizer.step()

            model.graph_memory.write_memory(new_nodes, new_edges)

            total_loss += loss.item()

            if step % 1 == 0:
                print(f"Epoch [{epoch+1}/{epochs}] | Step [{step}/{len(dataloader)}] | Loss: {loss.item():.4f}")
                
                ws[f"{kind}{next_row}"] = total_loss/1
                total_loss = 0
                next_row += 1
        

    wb.save(file_name)
"""
study(dataloader=trendy_sent,kind="A")
study(dataloader=it_sent,kind="C")
study(dataloader=bio_sent,kind="E")

study(dataloader=trendy_sent,kind="B")
study(dataloader=it_sent,kind="D")
study(dataloader=bio_sent,kind="F")
"""
data1 = ["나 아는사람 강다니엘 닮은 이모가 다시보게되는게 다시 그때처럼 안닮게 엄마보면 느껴지는걸수도 있는거임? 엄마도?"]
data2 = ["나랏말싸미 듕귁에 달아문자와로 서르 사맛디 아니할쎄 이런 전차로 어린 백셩이 니르고져 홇베이셔도 마참네 제 뜨들 시러펴디 몯핧 노미하니아 내 이랄 윙하야 어엿비너겨 새로 스믈 여듫 짜랄 맹가노니사람마다 해여 수비니겨 날로 쑤메 뻔한킈 하고져 할따라미니라"]

gang = pre_process.get_custom_dataloader(data1)
hoon = pre_process.get_custom_dataloader(data2)

study(dataloader=gang ,kind="M")
study(dataloader=hoon,kind="O")

study(dataloader=gang ,kind="N")
study(dataloader=hoon,kind="P")
